import openpyxl

wb = openpyxl.load_workbook('/home/ubuntu/upload/pasted_file_P2ivXm_小型样品点位标注图.xlsx')
ws = wb.active

print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")
print()

# 打印完整网格
print("=== 完整网格 ===")
for row in range(1, ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is None:
            v = ''
        vals.append(str(v).rjust(5))
    print(f"Row {row:2d}: {''.join(vals)}")

print()

# 解析点位映射: 找出每个编号(1-100)在16列网格中的位置
# 网格是 16行 x 16列 (但第一行可能是标题)
# 先看第一行
print("=== 第一行 ===")
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=col).value
    print(f"  Col {col}: {v} (type: {type(v).__name__})")

print()

# 构建点位映射: 编号 -> (row_in_grid, col_in_grid)
# 以及反向: (row_in_grid, col_in_grid) -> 编号
point_map = {}  # 编号 -> (grid_row, grid_col) 0-indexed

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None and isinstance(v, (int, float)):
            num = int(v)
            if 1 <= num <= 100:
                grid_row = row - 1  # 0-indexed
                grid_col = col - 1  # 0-indexed
                point_map[num] = (grid_row, grid_col)

print(f"Found {len(point_map)} points")
print()

# 按编号排序打印
print("=== 点位映射 (编号 -> 网格位置) ===")
for num in sorted(point_map.keys()):
    r, c = point_map[num]
    print(f"  Point {num:3d} -> grid[{r}][{c}]")

print()

# 分析数据帧中的顺序
# 原始数据是 256 字节，前 2 字节是 order+type，后面 128 字节是数据
# 两帧合并后 256 字节有效数据
# 编号 1-100 对应数据帧中的第几个字节？

# 从网格看:
# 第一行(row0): 列6-15 有编号 40,39,38,37,36,35,34,33,32,31
# 第二行(row1): 列6-15 有编号 50,49,48,47,46,45,44,43,42,41
# ...
# 所以编号是从右下角开始，先行后列

# 关键问题: 编号1-100在256字节数据中的索引是什么？
# 看起来编号就是数据帧中的字节索引(1-based)
# 即 point_n 对应 data[n-1]

# 生成 10x10 矩阵的映射数组
# 需要确定 10x10 矩阵的排列方式
# 从网格看，数据分布在 16x16 的网格中
# 上半部分(row 0-1): 编号 31-50 (2行 x 10列)
# 下半部分(row 8-15): 编号 1-30, 51-100 (8行 x 10列)

# 实际上看起来是两个区域:
# 上区: row 0-1, col 6-15 -> 编号 31-50 (从右到左，从上到下)
# 下区: row 8-15, col 6-15 -> 编号 1-30, 51-100

# 让我按行分组
print("=== 按行分组 ===")
row_groups = {}
for num in sorted(point_map.keys()):
    r, c = point_map[num]
    if r not in row_groups:
        row_groups[r] = []
    row_groups[r].append((num, c))

for r in sorted(row_groups.keys()):
    items = sorted(row_groups[r], key=lambda x: x[1])
    print(f"  Row {r}: {items}")

print()

# 生成 JavaScript 映射数组
# 10x10 矩阵: 按照网格中的物理位置排列
# 需要将 16x16 网格中的 100 个点映射到 10x10 矩阵

# 从网格看，有效区域是 col 6-15 (10列)
# 行的分布: row 0,1 (上区) 和 row 8-15 (下区) = 共10行
# 所以 10x10 矩阵的行对应: [0, 1, 8, 9, 10, 11, 12, 13, 14, 15]

matrix_rows = []
for r in sorted(row_groups.keys()):
    matrix_rows.append(r)

print(f"Matrix rows (grid rows with data): {matrix_rows}")
print(f"Total rows: {len(matrix_rows)}")

# 生成映射: matrix[i][j] = data_index (0-based)
# 其中 i 是 10x10 矩阵的行(0-9), j 是列(0-9)
print()
print("=== 10x10 矩阵映射 (data_index, 0-based) ===")
mapping = []
for i, grid_row in enumerate(matrix_rows):
    row_data = []
    for j in range(10):
        grid_col = j + 6  # 有效数据从第6列开始
        # 找这个位置对应的编号
        found = False
        for num, (r, c) in point_map.items():
            if r == grid_row and c == grid_col:
                row_data.append(num - 1)  # 0-based index
                found = True
                break
        if not found:
            row_data.append(-1)
    mapping.append(row_data)
    print(f"  Row {i}: {row_data}")

print()
print("=== JavaScript 映射数组 ===")
print("const smallSampleMap = [")
for i, row in enumerate(mapping):
    print(f"  {row},  // matrix row {i} -> grid row {matrix_rows[i]}")
print("];")
